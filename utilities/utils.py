import inspect
import softest
import logging
from openpyxl import workbook, load_workbook
import csv

class Utils(softest.TestCase):
    def assertlistitemtext(self, listofitems, value):
        for item in listofitems:
            print("The text is:" + item.text)
            self.soft_assert(self.assertEqual, item.text, value)
            if item.text == value:
                print("Test pass")
            else:
                print("Test failed")
        self.assert_all()

    def custom_logger(self, loglevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create a logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        # create console handler or file handler and set the log lover
        console_handler = logging.StreamHandler()
        # file_handler = logging.FileHandler(".//logs//automation.log")
        # create a formatter - how you want your logs to be formatted
        # file_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s --> %(message)s")
        console_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s --> %(message)s")
        # add formatter to file or console handler
        # file_handler.setFormatter(file_formatter)
        console_handler.setFormatter(console_formatter)
        # add console handler or file handler to logger
        # logger.addHandler(file_handler)
        logger.addHandler(console_handler)

        return logger


    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []

            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist


    def read_data_from_csv(file_name):
        datalist = []
        csvdata = open(file_name, "r")
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist